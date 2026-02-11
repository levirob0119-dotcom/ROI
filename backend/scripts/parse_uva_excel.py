#!/usr/bin/env python3
"""
解析 UVA模型数据底表.xlsx 文件，生成各车型的 UVA 矩阵 JSON 数据。
"""
import json
import os
import sys
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.join(SCRIPT_DIR, '..', '..')
EXCEL_PATH = os.path.join(PROJECT_ROOT, 'UVA模型数据底表.xlsx')
DATA_DIR = os.path.join(SCRIPT_DIR, '..', 'data')
MATRIX_DIR = os.path.join(DATA_DIR, 'uva-matrix')

# PETS ID 映射 (列序号 -> PETS key)
PETS_COLUMNS = {
    5: 'intelligent_driving',      # 智能驾驶
    6: 'intelligent_cockpit',      # 智能座舱
    7: 'safety',                   # 安全体验
    8: 'exterior_design',          # 外观设计及车身外部功能件
    9: 'interior_design',          # 内饰设计
    10: 'driving_experience',      # 驾驶体验
    11: 'riding_experience',       # 乘坐体验
    12: 'space',                   # 空间体验
    13: 'cabin_comfort',           # 座舱环境与舒适
    14: 'range_charging',          # 续航 & 补能体验
}

def parse_vehicle_sheet(ws, sheet_name):
    """解析一个车型的数据表"""
    matrix_data = []
    current_l1_name = None
    current_l1_category = None
    current_l1_weight = None
    
    # 跳过标题行(第1行)，从第2行开始读取数据
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        values = [cell.value for cell in row]
        
        # 列定义:
        # [0] UV L1, [1] L1 优先级(分类), [2] 目标UV L1权重, [3] UV L2, [4] 目标UV L2权重
        # [5-14] 10个PETS的分值
        
        l1_name = values[0]
        l1_category = values[1]
        l1_weight = values[2]
        l2_name = values[3]
        l2_weight = values[4]
        
        # 如果 L2 名称为空，跳过
        if not l2_name:
            continue
        
        # 更新 L1 信息（L1 名称只在每个分组的第一行出现）
        if l1_name:
            current_l1_name = str(l1_name).strip()
        if l1_category:
            current_l1_category = str(l1_category).strip()
        if l1_weight is not None:
            current_l1_weight = float(l1_weight) if l1_weight else 0
        
        # 构建 PETS 分值字典
        pets_scores = {}
        for col_idx, pets_key in PETS_COLUMNS.items():
            val = values[col_idx] if col_idx < len(values) else 0
            # 处理 None 和非数字值
            if val is None or val == '' or val == '#N/A':
                pets_scores[pets_key] = 0.0
            else:
                try:
                    pets_scores[pets_key] = round(float(val), 2)
                except (ValueError, TypeError):
                    pets_scores[pets_key] = 0.0
        
        entry = {
            "l1_name": current_l1_name,
            "l1_category": current_l1_category,
            "l1_weight": round(current_l1_weight, 4) if current_l1_weight else 0,
            "l2_name": str(l2_name).strip(),
            "l2_weight": round(float(l2_weight), 2) if l2_weight else 0,
            "pets_scores": pets_scores
        }
        
        matrix_data.append(entry)
    
    return matrix_data


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ 文件不存在: {EXCEL_PATH}")
        sys.exit(1)
    
    print(f"📖 正在读取 Excel 文件: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, data_only=True)
    
    print(f"📋 工作表列表: {wb.sheetnames}")
    
    # 确保输出目录存在
    os.makedirs(MATRIX_DIR, exist_ok=True)
    
    # 查找所有车型数据底表
    vehicle_sheets = {}
    for sheet_name in wb.sheetnames:
        if '数据底表' in sheet_name:
            # 从 sheet 名中提取车型名称
            vehicle_name = sheet_name.replace('数据底表', '').strip()
            vehicle_sheets[vehicle_name.lower()] = sheet_name
    
    if not vehicle_sheets:
        print("⚠️ 没有找到包含 '数据底表' 的工作表")
        sys.exit(1)
    
    print(f"\n🚗 找到 {len(vehicle_sheets)} 个车型数据:")
    for vehicle_id, sheet_name in vehicle_sheets.items():
        print(f"   {vehicle_id} -> {sheet_name}")
    
    # 解析并保存每个车型数据
    for vehicle_id, sheet_name in vehicle_sheets.items():
        ws = wb[sheet_name]
        print(f"\n📊 正在解析: {sheet_name}")
        
        matrix_data = parse_vehicle_sheet(ws, sheet_name)
        
        if not matrix_data:
            print(f"   ⚠️ 没有解析到数据")
            continue
        
        # 统计信息
        l1_names = set(item['l1_name'] for item in matrix_data)
        categories = set(item['l1_category'] for item in matrix_data)
        
        print(f"   ✅ 解析到 {len(matrix_data)} 条 L2 数据")
        print(f"   📁 L1 分类数: {len(l1_names)}")
        print(f"   📁 需求分类: {categories}")
        
        # 保存 JSON 文件
        output_path = os.path.join(MATRIX_DIR, f"{vehicle_id}.json")
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(matrix_data, f, ensure_ascii=False, indent=2)
        
        print(f"   💾 已保存到: {output_path}")
        
        # 打印前3条数据作为验证
        print(f"   📋 前3条数据预览:")
        for i, item in enumerate(matrix_data[:3]):
            print(f"      [{i+1}] {item['l1_name']} / {item['l2_name']} (权重: {item['l2_weight']})")
            scores_preview = {k: v for k, v in item['pets_scores'].items() if v > 0}
            print(f"           PETS分值(非零): {scores_preview}")
    
    print(f"\n✅ 所有车型数据已更新完成！")


if __name__ == '__main__':
    main()
